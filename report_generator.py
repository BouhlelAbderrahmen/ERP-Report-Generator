import os
import sys
import time
import pandas as pd
import numpy as np
import yaml
import calendar
import warnings
import logging
from datetime import datetime
from sqlalchemy import create_engine, text
from sqlalchemy.exc import OperationalError
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from fpdf import FPDF
from dotenv import load_dotenv

# 1. SETUP & LOGGING
warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

load_dotenv() 
OUTPUT_DIR = "/app/outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

print("=== STARTING OPTIMIZED REPORT GENERATION ===")

# 2. CONFIGURATION & PARAMS
try:
    with open("/app/config.yaml", "r") as f:
        config = yaml.safe_load(f)
    print("✔ Configuration loaded successfully")
except FileNotFoundError:
    print("CRITICAL: config.yaml not found!")
    sys.exit(1)

DB_USER = os.getenv("DB_USER")
DB_PASS = os.getenv("DB_PASSWORD")
DB_HOST = os.getenv("DB_HOST")
DB_NAME = os.getenv("DB_NAME")
DB_PORT = int(os.getenv("DB_PORT", "3306"))

TARGET_MONTH = os.getenv("REPORT_MONTH", "09")
TARGET_YEAR = os.getenv("REPORT_YEAR", "2024")

VAT_RATE = config['business_rules']['vat_rate']
EXCLUDED = config['business_rules']['excluded_families']
FALLBACK_MARGIN = config['business_rules']['margin_fallback']
TOP_N = config['report_settings']['top_n_stores']
vat_divisor = 1 + (VAT_RATE / 100)
start_time = datetime.now()

# 3. DATABASE CONNECTION
def connect_to_database(max_retries=60, retry_interval=10):
    connection_string = f"mysql+mysqlconnector://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
    print(f"Connecting to {DB_HOST}... (Waiting for 7GB import to finish)")
    for attempt in range(1, max_retries + 1):
        try:
            engine = create_engine(connection_string, pool_pre_ping=True)
            with engine.connect() as conn:
                conn.execute(text("SELECT 1 FROM histovente LIMIT 1"))
                print(f"✔ Database ready after attempt {attempt}")
                return engine
        except Exception:
            if attempt % 5 == 0:
                print(f"⏳ Still waiting for database... (Attempt {attempt}/{max_retries})")
            time.sleep(retry_interval)
    print("CRITICAL: Database timed out.")
    sys.exit(1)

engine = connect_to_database()

# 4. DATE PREPARATION
last_day = calendar.monthrange(int(TARGET_YEAR), int(TARGET_MONTH))[1]
start_date = f"{TARGET_YEAR}-{TARGET_MONTH}-01"
end_date = f"{TARGET_YEAR}-{TARGET_MONTH}-{last_day}"

# 5. OPTIMIZED SQL QUERY (Prevents Cartesian Product/Duplicates)
excluded_str = "', '".join(EXCLUDED)
sql_query = f"""
SELECT 
    h.CodeMag AS StoreCode, 
    h.Quantite AS Quantity, 
    h.total AS TotalTTC,
    ROUND(h.total / {vat_divisor}, 2) AS TotalHT,
    (a.PrixAchat * h.Quantite) AS PurchasePrice
FROM histovente h
LEFT JOIN (
    /* DUPLICATE PREVENTION: Use MAX(IdEntite) to ensure 1 Barcode = 1 Article */
    SELECT CodeBarre, MAX(IdEntite) as IdEntite 
    FROM codebarre 
    GROUP BY CodeBarre
) c ON c.CodeBarre = h.Barcode 
LEFT JOIN article a ON a.IDArticle = c.IdEntite 
WHERE h.reception BETWEEN '{start_date}' AND '{end_date}'
AND h.typevente IN ('Avoir', 'Vente')
AND h.Famille NOT IN ('{excluded_str}');
"""

print(f"🚀 Pulling and Processing {TARGET_MONTH}/{TARGET_YEAR} data...")
df_sql = pd.read_sql(sql_query, engine)

# Validation Point
raw_total = df_sql['TotalHT'].sum()
print(f"📊 DATA CHECK: Total Revenue HT in DB: {raw_total:,.2f}")

# 6. AGGREGATION
ca = df_sql.groupby(['StoreCode'])[['Quantity', 'TotalHT', 'PurchasePrice']].sum().reset_index()

# 7. DATA ENRICHMENT (Excel & CSV Merges)
print("--- Data Enrichment Phase ---")

# Surface Area Merge
surface_file = '/app/data/surface_magasin.xlsx'
try:
    if os.path.exists(surface_file):
        df_surface = pd.read_excel(surface_file)
        # Ensure StoreCodes are strings for a clean match
        ca['StoreCode'] = ca['StoreCode'].astype(str)
        df_surface['Code'] = df_surface['Code'].astype(str)
        
        ca = pd.merge(ca, df_surface[['Code', 'superficie M2']], 
                     left_on='StoreCode', right_on='Code', how='left')
        ca['Surface'] = ca['superficie M2'].fillna(0)
        print(f"✔ Merged with Surface data")
    else:
        print(f"⚠ Surface file not found")
        ca['Surface'] = 0
except Exception as e:
    print(f"⚠ Surface error: {e}")
    ca['Surface'] = 0

# BW Reference Merge
bw_file = '/app/data/BWSales09-2024.csv'
try:
    if os.path.exists(bw_file):
        df_bw = pd.read_csv(bw_file, sep=',', quotechar='"')
        df_bw['StoreCode'] = df_bw['StoreCode'].astype(str)
        
        ca = pd.merge(ca, df_bw[['StoreCode', 'TotalHT']], 
                     on='StoreCode', how='left', suffixes=('', '_BW'))
        ca['BW_Turnover'] = ca['TotalHT_BW'].fillna(0)
        print(f"✔ Merged with BW Reference")
    else:
        ca['BW_Turnover'] = 0
except Exception as e:
    print(f"⚠ BW error: {e}")
    ca['BW_Turnover'] = 0

# 8. KPI CALCULATIONS
# Fallback logic if PurchasePrice is missing or zero
ca['PurchasePrice'] = ca['PurchasePrice'].replace(0, np.nan)
ca['PurchasePrice'] = ca['PurchasePrice'].fillna(ca['TotalHT'] * (1 - FALLBACK_MARGIN))

ca['Profit'] = (ca['TotalHT'] - ca['PurchasePrice']).round(2)
ca['GP %'] = ((ca['Profit'] / ca['TotalHT']) * 100).replace([np.inf, -np.inf], 0).fillna(0).round(2)
ca['Yield_m2'] = np.where(ca['Surface'] > 0, (ca['TotalHT'] / ca['Surface']).round(2), 0)
ca['Gap_%'] = np.where(ca['BW_Turnover'] > 0, (((ca['TotalHT'] / ca['BW_Turnover']) - 1) * 100).round(2), 0)

# 9. OUTPUT GENERATION
print("\n--- Saving Results ---")
excel_file = os.path.join(OUTPUT_DIR, f'Sales_Analysis_{TARGET_YEAR}_{TARGET_MONTH}.xlsx')
try:
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        ca.sort_values('Yield_m2', ascending=False).to_excel(writer, sheet_name='Summary', index=False)
        # We only save a sample of raw data to keep Excel from crashing
        df_sql.head(500000).to_excel(writer, sheet_name='Sample_Raw_Data', index=False)
    print(f"✔ Excel Saved: {excel_file}")
except Exception as e:
    print(f"✗ Excel Error: {e}")

# PDF Executive Summary
class PDF(FPDF):
    def header(self):
        self.set_fill_color(40, 40, 40)
        self.rect(0, 0, 210, 40, 'F')
        self.set_text_color(255, 255, 255)
        self.set_font('helvetica', 'B', 18)
        self.cell(0, 20, 'NAF NAF ADVANCED ANALYTICS', ln=True, align='C')
        self.ln(10)

try:
    pdf = PDF()
    pdf.add_page()
    pdf.set_text_color(0)
    pdf.set_font("helvetica", 'B', 10)
    pdf.cell(0, 10, f"Executive Report: {TARGET_MONTH}/{TARGET_YEAR}", ln=True)
    
    pdf.set_fill_color(240, 240, 240)
    cols = ["Store", "Surf(m2)", "Rev HT", "Yield/m2", "Margin%", "BW Gap%"]
    for col in cols:
        pdf.cell(32, 10, col, 1, 0, 'C', True)
    pdf.ln()

    pdf.set_font("helvetica", '', 9)
    for _, row in ca.sort_values('Yield_m2', ascending=False).head(TOP_N).iterrows():
        pdf.cell(32, 8, str(row['StoreCode']), 1)
        pdf.cell(32, 8, f"{row['Surface']:.1f}", 1, 0, 'C')
        pdf.cell(32, 8, f"{row['TotalHT']:,.0f}", 1, 0, 'R')
        pdf.cell(32, 8, f"{row['Yield_m2']:.2f}", 1, 0, 'R')
        pdf.cell(32, 8, f"{row['GP %']}%", 1, 0, 'R')
        pdf.cell(32, 8, f"{row['Gap_%']}%", 1, 1, 'R')

    pdf_path = os.path.join(OUTPUT_DIR, f'Executive_Summary_{TARGET_YEAR}_{TARGET_MONTH}.pdf')
    pdf.output(pdf_path)
    print(f"✔ PDF Saved: {pdf_path}")
except Exception as e:
    print(f"✗ PDF Error: {e}")

engine.dispose()
print(f"\n--- FINISH: Runtime: {datetime.now() - start_time} ---")